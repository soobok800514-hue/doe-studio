"""
Excel / PDF 내보내기 모듈
"""
import io
from datetime import datetime
import pandas as pd
import numpy as np


# ─────────────────────────────────────────────────────────
# Excel 내보내기  (openpyxl — 이미 설치됨)
# ─────────────────────────────────────────────────────────

def to_excel(
    df_data: pd.DataFrame,
    mean_tbl: pd.DataFrame,
    sn_tbl: pd.DataFrame,
    anova_df: pd.DataFrame,
    opt_df: pd.DataFrame,
    y_pred: float,
    response_col: str,
    sn_type: str,
    factor_cols: list,
) -> bytes:
    """분석 결과를 5-시트 Excel 파일로 생성 → bytes 반환."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 공통 스타일
    BLUE_FILL  = PatternFill("solid", fgColor="1F4E79")
    GRAY_FILL  = PatternFill("solid", fgColor="D9E1F2")
    PRED_FILL  = PatternFill("solid", fgColor="FFF2CC")
    WHITE_FONT = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=10)
    DARK_FONT  = Font(name="맑은 고딕", bold=True, size=10)
    BODY_FONT  = Font(name="맑은 고딕", size=10)
    TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
    SEC_FONT   = Font(name="맑은 고딕", bold=True, size=11, color="1F4E79")
    RED_FONT   = Font(name="맑은 고딕", bold=True, size=12, color="C00000")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _s         = Side(style="thin", color="BBBBBB")
    BORDER     = Border(left=_s, right=_s, top=_s, bottom=_s)
    now_str    = datetime.now().strftime("%Y-%m-%d %H:%M")

    def _cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def _header(ws, row, cols, sc=1):
        for c, val in enumerate(cols, sc):
            cell = ws.cell(row, c, str(val))
            cell.font, cell.fill = WHITE_FONT, BLUE_FILL
            cell.alignment, cell.border = CENTER, BORDER

    def _cell(ws, row, col, val, font=None, fill=None):
        cell = ws.cell(row, col)
        try:
            cell.value = round(float(val), 4) if pd.notnull(val) else ""
        except Exception:
            cell.value = str(val) if val is not None else ""
        cell.font = font or BODY_FONT
        if fill:
            cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER
        return cell

    def _write_response_tbl(ws, df, start_row, title):
        ws.cell(start_row, 1, title).font = SEC_FONT
        r = start_row + 1
        header = [""] + [str(c) for c in df.columns]
        _header(ws, r, header)
        r += 1
        for idx, row_data in df.iterrows():
            c0 = ws.cell(r, 1, str(idx))
            c0.font, c0.fill = DARK_FONT, GRAY_FILL
            c0.alignment, c0.border = CENTER, BORDER
            for c, val in enumerate(row_data, 2):
                _cell(ws, r, c, val)
            r += 1
        for c in range(1, len(header) + 1):
            _cw(ws, c, 14)
        return r + 1

    # ── Sheet 1: 분석 개요 ───────────────────────────────
    ws0 = wb.create_sheet("분석 개요")
    ws0.sheet_view.showGridLines = False
    ws0['B2'] = "DOE Studio — Taguchi 분석 결과 보고서"
    ws0['B2'].font = TITLE_FONT
    ws0['B3'] = f"생성일시: {now_str}"
    ws0['B3'].font = Font(name="맑은 고딕", size=10, color="888888")
    info = [
        ("응답 변수", response_col),
        ("인자",      ", ".join(factor_cols)),
        ("S/N 비 유형", {"larger": "망대(Larger)", "smaller": "망소(Smaller)", "nominal": "망목(Nominal)"}.get(sn_type, sn_type)),
        ("총 시험 횟수", f"{len(df_data)}회"),
        ("예측 최적 응답값", f"{y_pred:.4f}"),
    ]
    for i, (k, v) in enumerate(info, 5):
        _cell(ws0, i, 2, k, font=DARK_FONT, fill=GRAY_FILL)
        _cell(ws0, i, 3, v)
    _cw(ws0, 2, 22); _cw(ws0, 3, 40)

    # ── Sheet 2: 원본 데이터 ─────────────────────────────
    ws1 = wb.create_sheet("원본 데이터")
    ws1.sheet_view.showGridLines = False
    ws1['A1'] = "원본 데이터"; ws1['A1'].font = SEC_FONT
    cols = list(df_data.columns)
    _header(ws1, 2, cols)
    for i, (_, row_data) in enumerate(df_data.iterrows(), 3):
        for c, val in enumerate(row_data, 1):
            _cell(ws1, i, c, val)
    for c, col in enumerate(cols, 1):
        _cw(ws1, c, max(12, len(str(col)) + 4))

    # ── Sheet 3: 응답표 ──────────────────────────────────
    ws2 = wb.create_sheet("응답표")
    ws2.sheet_view.showGridLines = False
    r = 1
    if mean_tbl is not None:
        r = _write_response_tbl(ws2, mean_tbl, r, "평균 응답표 (Response Table for Means)")
    if sn_tbl is not None:
        r = _write_response_tbl(ws2, sn_tbl, r, f"S/N 응답표 ({sn_type})")

    # ── Sheet 4: ANOVA ───────────────────────────────────
    ws3 = wb.create_sheet("ANOVA")
    ws3.sheet_view.showGridLines = False
    ws3['A1'] = "ANOVA (분산분석)"; ws3['A1'].font = SEC_FONT
    if anova_df is not None:
        cols = list(anova_df.columns)
        _header(ws3, 2, cols)
        for i, (_, row_data) in enumerate(anova_df.iterrows(), 3):
            for c, val in enumerate(row_data, 1):
                _cell(ws3, i, c, val)
        for c, col in enumerate(cols, 1):
            _cw(ws3, c, max(15, len(str(col)) + 4))

    # ── Sheet 5: 최적 수준 ───────────────────────────────
    ws4 = wb.create_sheet("최적 수준")
    ws4.sheet_view.showGridLines = False
    ws4['A1'] = "최적 수준 및 예측값"; ws4['A1'].font = SEC_FONT
    if opt_df is not None:
        cols = list(opt_df.columns)
        _header(ws4, 2, cols)
        for i, (_, row_data) in enumerate(opt_df.iterrows(), 3):
            for c, val in enumerate(row_data, 1):
                _cell(ws4, i, c, val)
        for c, col in enumerate(cols, 1):
            _cw(ws4, c, max(16, len(str(col)) + 4))
        r_p = len(opt_df) + 4
        _cell(ws4, r_p, 1, "예측 최적 응답값", font=DARK_FONT, fill=PRED_FILL)
        _cell(ws4, r_p, 2, y_pred, font=RED_FONT, fill=PRED_FILL)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────
# PDF 내보내기  (reportlab — 한글 CID 폰트 사용)
# ─────────────────────────────────────────────────────────

def to_pdf(
    df_data: pd.DataFrame,
    mean_tbl: pd.DataFrame,
    sn_tbl: pd.DataFrame,
    anova_df: pd.DataFrame,
    opt_df: pd.DataFrame,
    y_pred: float,
    response_col: str,
    sn_type: str,
    factor_cols: list,
) -> bytes:
    """분석 결과를 한글 PDF 보고서로 생성 → bytes 반환."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # 한글 CID 폰트 (Adobe Reader 내장, 설치 불필요)
    pdfmetrics.registerFont(UnicodeCIDFont('HYGothic-Medium'))
    KOR = 'HYGothic-Medium'

    BLUE  = colors.HexColor("#1F4E79")
    LGRAY = colors.HexColor("#D9E1F2")
    YGRAY = colors.HexColor("#FFF2CC")
    RED   = colors.HexColor("#C00000")
    WHITE = colors.white

    title_st = ParagraphStyle("t", fontName=KOR, fontSize=15, textColor=BLUE,
                               spaceAfter=3, leading=20)
    h2_st    = ParagraphStyle("h2", fontName=KOR, fontSize=11, textColor=BLUE,
                               spaceAfter=3, leading=15, spaceBefore=10)
    small_st = ParagraphStyle("sm", fontName=KOR, fontSize=8,
                               textColor=colors.grey, leading=11)

    def _tbl_style():
        return TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  BLUE),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",       (0, 0), (-1, -1), KOR),
            ("FONTSIZE",       (0, 0), (-1, -1), 8),
            ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LGRAY]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 4),
            ("TOPPADDING",     (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
        ])

    def _is_num(v):
        try:
            float(v); return True
        except Exception:
            return False

    def _fmt(v):
        try:
            return f"{float(v):.4f}" if _is_num(v) else str(v)
        except Exception:
            return str(v)

    def _df_table(df, idx_col=True):
        if idx_col:
            header = [""] + [str(c) for c in df.columns]
            rows   = [header] + [
                [str(i)] + [_fmt(v) for v in row]
                for i, row in zip(df.index, df.values)
            ]
        else:
            header = [str(c) for c in df.columns]
            rows   = [header] + [[_fmt(v) for v in row] for row in df.values]
        W   = 17 * cm
        cw  = [W / len(header)] * len(header)
        tbl = Table(rows, colWidths=cw, repeatRows=1)
        tbl.setStyle(_tbl_style())
        return tbl

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm,  bottomMargin=2*cm)
    story = []

    # 제목
    story.append(Paragraph("DOE Studio — Taguchi 분석 결과 보고서", title_st))
    story.append(Paragraph(f"생성일시: {now_str}", small_st))
    story.append(HRFlowable(width="100%", thickness=1.5, color=BLUE, spaceAfter=6))

    # 분석 개요
    story.append(Paragraph("1. 분석 개요", h2_st))
    sn_label = {"larger": "망대 (Larger-the-better)",
                 "smaller": "망소 (Smaller-the-better)",
                 "nominal": "망목 (Nominal-the-best)"}.get(sn_type, sn_type)
    info_rows = [
        ["항목", "내용"],
        ["응답 변수", response_col],
        ["인자", ", ".join(factor_cols)],
        ["S/N 비 유형", sn_label],
        ["총 시험 횟수", f"{len(df_data)}회"],
        ["예측 최적 응답값", f"{y_pred:.4f}"],
    ]
    info_tbl = Table(info_rows, colWidths=[5*cm, 12*cm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  BLUE),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",       (0, 0), (-1, -1), KOR),
        ("FONTSIZE",       (0, 0), (-1, -1), 9),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LGRAY]),
        ("LEFTPADDING",    (0, 0), (-1, -1), 6),
        ("TOPPADDING",     (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
        ("BACKGROUND",     (0, 5), (-1, 5),  YGRAY),
        ("FONTSIZE",       (1, 5), (1, 5),   10),
        ("TEXTCOLOR",      (1, 5), (1, 5),   RED),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.3*cm))

    # 원본 데이터
    story.append(Paragraph("2. 원본 데이터", h2_st))
    cols  = list(df_data.columns)
    drows = [cols] + [[_fmt(v) for v in row] for _, row in df_data.iterrows()]
    dtbl  = Table(drows, colWidths=[17*cm/len(cols)]*len(cols), repeatRows=1)
    dtbl.setStyle(_tbl_style())
    story.append(dtbl)
    story.append(Spacer(1, 0.3*cm))

    # 응답표
    if mean_tbl is not None:
        story.append(Paragraph("3. 평균 응답표 (Response Table for Means)", h2_st))
        story.append(_df_table(mean_tbl))
        story.append(Spacer(1, 0.2*cm))

    if sn_tbl is not None:
        story.append(Paragraph(f"4. S/N 응답표 ({sn_type})", h2_st))
        story.append(_df_table(sn_tbl))
        story.append(Spacer(1, 0.2*cm))

    # ANOVA
    if anova_df is not None:
        story.append(Paragraph("5. ANOVA (분산분석)", h2_st))
        story.append(_df_table(anova_df, idx_col=False))
        story.append(Spacer(1, 0.2*cm))

    # 최적 수준
    if opt_df is not None:
        story.append(Paragraph("6. 최적 수준 및 예측값", h2_st))
        story.append(_df_table(opt_df, idx_col=False))
        story.append(Spacer(1, 0.2*cm))

        pred_rows = [["예측 최적 응답값", f"{y_pred:.4f}"]]
        pred_tbl  = Table(pred_rows, colWidths=[8.5*cm, 8.5*cm])
        pred_tbl.setStyle(TableStyle([
            ("FONTNAME",       (0, 0), (-1, -1), KOR),
            ("FONTSIZE",       (0, 0), (0, 0),   10),
            ("FONTSIZE",       (1, 0), (1, 0),   13),
            ("TEXTCOLOR",      (1, 0), (1, 0),   RED),
            ("BACKGROUND",     (0, 0), (-1, -1), YGRAY),
            ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#BBBBBB")),
            ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING",     (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
        ]))
        story.append(pred_tbl)

    # 푸터
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        "본 보고서는 DOE Studio (사내 Taguchi 분석 자동화 도구)에서 자동 생성되었습니다.",
        small_st,
    ))

    doc.build(story)
    return buf.getvalue()
